import json
from sympy import *
import numpy as np
from matplotlib import pyplot as plt
from matplotlib.colors import rgb2hex
# from sklearn import cluster
import math
import random
import string
import csv
import xlrd
import xlwt
from openpyxl import load_workbook


def load_data(path):
    # 1.工序的加工时间，约束，工厂位置数据，质量
    with open(path, 'r') as file:
        data = json.load(file)
    # 2.装配中心位于原点
    # print(len(data))
    car_position = [(0, 0)]
    # print('load data over！')
    return data, car_position


# 距离公式计算
def calculate(x1, y1, x2, y2):
    return round(math.sqrt(math.pow(x1-x2, 2)+math.pow(y1-y2, 2)), 2)


# 数据预处理
def data_preprocessing(data, num_station):
    car_maximizes_capacity = 40  # 每辆车载重
    cat_v = 2  # 车辆的速度设为定置 米/s
    # get仓库位置
    position_warehouse = []
    for i in range(len(data)):
        position_warehouse.append(data[i][3][0])

    # get 所有工件的总重量和所用最小车辆数量
    m_op = 0
    for i in range(len(data)):
        m_op += data[i][4][0]
    # print('m_op: '+str(m_op))
    # 所用最小车辆数量
    n_car = math.ceil(m_op/car_maximizes_capacity)   # 避免装不下

    # calculation 所有两点之间的距离 & 折算为时间
    warehouse_to_house_distance = []
    for i in range(len(data)):
        warehouse_to_house_distance.append([])
        for j in range(len(data)):
            warehouse_to_house_distance[i].append([])
    for i in range(len(data)):
        for j in range(len(data)):
            warehouse_to_house_distance[i][j] = calculate(position_warehouse[i][0], position_warehouse[i][1],
                                                          position_warehouse[j][0], position_warehouse[j][1])
    wh_to_wh_time = []
    for i in range(len(warehouse_to_house_distance)):
        wh_to_wh_time.append([])
        for j in range(len(warehouse_to_house_distance)):
            wh_to_wh_time[i].append([])
    for i in range(len(warehouse_to_house_distance)):
        for j in range(len(warehouse_to_house_distance)):
            wh_to_wh_time[i][j] = (round((warehouse_to_house_distance[i][j]/cat_v), 6))

    # calculate 仓库到中心的距离 & 折算成时间
    warehouse_to_center_distance = []
    for i in range(len(data)):
        warehouse_to_center_distance.append([])
    for i in range(len(data)):
        warehouse_to_center_distance[i] = calculate(position_warehouse[i][0], position_warehouse[i][1], 0, 0)
    # print(warehouse_to_center_distance)
    wh_to_center_time = []
    for i in range(len(warehouse_to_center_distance)):
        wh_to_center_time.append(round((warehouse_to_center_distance[i]/cat_v), 6))

    # 计算CM
    time_op_total = 0
    for i in range(len(data)):
        time_op_total += data[i][1]
    cm = round(time_op_total/num_station, 2)

    # 计算工序的前序有哪些
    pre_op = []
    for i in range(len(data)):
        pre_op.append([])
    for i in range(len(data)):
        if data[i][2]:  # 如果存在后序
            for j in range(len(data[i][2])):  # 后序的个数
                pre_op[data[i][2][j]-1].append(i+1)

    # 计算所有配件总重量
    total_op_m = 0
    for i in range(0, len(data)):
        total_op_m += round(data[i][4][0], 2)
    # print(total_op_m)

    return cat_v, n_car, position_warehouse, wh_to_wh_time, wh_to_center_time, cm, pre_op, car_maximizes_capacity, total_op_m, m_op


# 运输阶段
def transport(label, n_car, wh_to_wh, wh_to_center):
    car_tasks = []
    car_finish_time = []
    for i in range(n_car):
        car_tasks.append([])
    for i in range(len(label)):
        car_tasks[label[i]].append(i)

    car_finish_time = []
    for i in range(n_car):
        car_finish_time.append([])
        temp = 0
        for j in range(len(car_tasks[i])-1):
            temp += (wh_to_wh[car_tasks[i][j]][car_tasks[i][j+1]])
        car_finish_time[i] = round(temp, 2)

    for i in range(n_car):
        try:
            last_wh = len(car_tasks[i])
            car_finish_time[i] += (wh_to_center[car_tasks[i][0]] + wh_to_center[car_tasks[i][last_wh-1]])
            car_finish_time[i] = round(car_finish_time[i], 2)
        except IndexError:
            print(i)
            print(car_tasks)
            print(car_tasks[i][0])
            print(wh_to_center[car_tasks[i][0]])
            print(wh_to_center[car_tasks[i][last_wh-1]])

    # 每个工件抵达的时间
    arrive_time = []
    for i in range(len(label)):
        arrive_time.append([])
    for i in range(len(car_tasks)):
        for j in range(len(car_tasks[i])):
            arrive_time[car_tasks[i][j]] = car_finish_time[i]

    total_t = 0
    for i in range(len(arrive_time)):
        total_t += arrive_time[i]

    b = Symbol('b')
    b = solve([b*n_car/((1-b)*total_t) - 7/3], b)
    b = list(b.values())[0]
    cost = round(b*n_car + (1-b)*total_t, 2)

    return arrive_time, cost


# 装配阶段数据预处理
def data_prep_alb(first_op, ar_time, data):
    # 首个开始装配的工件的抵达时间
    first_alb = ar_time[first_op]
    wt_time = []
    for i in range(len(ar_time)):
        wt_time.append([])
    # 计算每个工序相对第一个
    for i in range(len(ar_time)):
        if (ar_time[i] - first_alb) <= 0:    # 来的早的工件要等着，所以等待时间设置为0
            wt_time[i] = 0
        else:
            wt_time[i] = round(ar_time[i] - first_alb, 6)

    time_op_total = 0
    for i in range(len(data)):
        time_op_total += data[i][1]

    return wt_time


def evaluate(n_car, arrive_time, ct):
    # 参数确定
    # a, b = 0.3, 0.999
    total_t = 0
    for i in range(len(arrive_time)):
        total_t += arrive_time[i]
    # 计算参数(计算一次)
    b = Symbol('b')
    a = Symbol('a')
    b = solve([b*n_car/((1-b)*total_t) - 7/3], b)
    b = list(b.values())[0]
    a = solve([(a*(b*n_car + (1-b)*total_t))/(1-a)*ct - 3/7], a)
    a = list(a.values())[0]
    # 评价函数公式
    cost = a*(b*n_car + (1-b)*total_t) + (1-a)*ct
    return cost


# 新建一个excel表，名字为数据名
def create_xls(data_name):
    excel_path = data_name+'.xls'
    work_book = xlwt.Workbook(encoding='utf-8')
    sheet = work_book.add_sheet(data_name)
    sheet.write(0, 0, '车辆数')
    sheet.write(0, 1, '工作站数')
    sheet.write(0, 2, '运输成本')
    sheet.write(0, 3, '节拍值')
    sheet.write(0, 4, '理想节拍')
    work_book.save(excel_path)


# 保存数据到建立的数据表
def save_data(n_car, num_station, trans_cost, this_ct, cm, num_data, data_name):
    # 读取Excel文件
    excel_path = data_name + r'.xls'
    work_book = xlwt.Workbook(encoding='utf-8')
    workbook = work_book.add_sheet(data_name)
    i = num_data
    workbook.write(i, 1, str(num_station))
    workbook.write(i, 1, str(num_station))
    workbook.write(i, 2, str(trans_cost))
    workbook.write(i, 3, str(this_ct))
    workbook.write(i, 4, str(cm))
    workbook.save(excel_path)  # 保存


